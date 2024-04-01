from datetime import datetime
import tkinter as tk
import traceback
from PIL import Image, ImageTk
from tkinter import filedialog, messagebox
import csv
import os
import random
import openpyxl

class FileOperationsApp:
    def __init__(self, root):
        self.root = root
        self.root.resizable(False, False)  # 禁止窗口的大小调整
        self.root.title("菜叶子数据处理小助手V3.0")
        self.root.geometry("650x420+750+200")
        self.root.iconbitmap('./pics/caiGou.ico')

        self.frame = 0
        self.frames = []

        self.label = tk.Label(root)
        self.label.pack()
        self.load_gif_frames()
        # self.update_animation()

        # 显示存在多个单位的商品名称的框体
        self.show_multi_unit_frame()

        button_props = {
            "height": 4,
            "width": 15,
            "bg": "#eacdd1",
            "font": ("Helvetica", 10)
        }
        button_texts = [
            ('月度\n数据汇总', self.delMonth_1, 0.15, 0.65),
            ('每日\n数据处理', self.delDay, 0.65, 0.65),
            # ('月度数据\n(最后归并)汇总', self.delMonth_2, 0.15, 0.75),
            # ('排查同商品\n存在多种单位', self.checkRepeat, 0.58, 0.75)
        ]

        for text, command_func, relx, rely in button_texts:
            tk.Button(self.root, text=text, command=command_func, **button_props).place(relx=relx, rely=rely)
        tk.Button(self.root, text='换个背景', command=self.load_gif_frames, height=2, width=10, bg="#8abcd1",
                  font=("Helvetica", 10)).place(relx=0.43, rely=0.7)

        current_time = datetime.now()
        formatted_time = current_time.strftime("%Y-%m-%d %H:%M:%S")
        footer_text = f" {formatted_time}  版权所有 © S in {current_time.year} "

        # 创建 footer_label
        self.footer_label = tk.Label(self.root, text="", fg="black", font=("Helvetica", 9), pady=5)
        self.footer_label.place(relx=0.5, rely=0.95, anchor="center", relwidth=0.99)
        self.update_app()

    def update_app(self):
        # 更新动画
        self.update_animation()
        # 更新footer_label的时间
        self.update_footer_time()
        # 调度下一个更新
        self.root.after(200, self.update_app)
    def update_footer_time(self):
        # 获取当前时间
        current_time = datetime.now()
        formatted_time = current_time.strftime("%Y-%m-%d %H:%M:%S")
        footer_text = f" {formatted_time}   -  版权所有 © S  -   In  {current_time.year} "

        # 更新 footer_label 的文本
        self.footer_label.config(text=footer_text)

    def load_gif_frames(self):
        random_gif = get_random_gif('./pics/')
        gif = Image.open('./pics/'+random_gif)
        gif_frames = []
        try:
            while True:
                gif_frames.append(gif.copy())
                gif.seek(len(gif_frames))
        except EOFError as e:
            pass


        self.frames = [ImageTk.PhotoImage(frame) for frame in gif_frames]
        self.frame = 0

    def update_animation(self):
        self.label.configure(image=self.frames[self.frame])
        self.frame = (self.frame + 1) % len(self.frames)

    def delMonth_1(self):
        file_path = filedialog.askopenfilename(filetypes=[('Excel Files', '*.xlsx;*.xls')])
        if not file_path:
            messagebox.showerror(title="通知", message="未获取到excel文件")
            return
        workbook = openpyxl.load_workbook(file_path)
        if '配送明细' not in workbook.sheetnames:
            messagebox.showerror(title="通知", message="没找到名称为【配送明细】的sheet表\n\n请注意将文件中数据文件所在sheet页面名称修改为“配送明细”才可以被识别")
            return None
        sheet = workbook['配送明细']
        commodity_dict = parse_data(sheet) # commodity_dict = {"橘子" : [["斤", 2, 2.6, "规格1"], ['箱', 13, 30, "规格2"]]}
        write_to_excel(file_path, commodity_dict)

    def delMonth_2(self):
        pass
    def delDay(self):
        try:
            file_path = filedialog.askopenfilename(filetypes=[('Excel Files', '*.xlsx;*.xls')])
            if not file_path:
                messagebox.showerror(title="错误！", message="未获取到excel文件")
                return
            process_excel(file_path)
        except Exception as e:
            messagebox.showerror(title="错误！", message=f"发生错误：{str(e)}")
            traceback.print_exc()
    def checkRepeat(self):
        pass

    def show_multi_unit_frame(self):
        lf = tk.LabelFrame(self.root, text='排查存在多个单位的商品名称结果显示处')
        lf.place(relx=0.1, rely=0.1, relwidth=0.8, relheight=0.4)
        self.ent = tk.Label(lf, bg='white', font=("黑体", 12), text="已经进行了优化\n\n已经不需要排查存在多个单位的商品名称")
        self.ent.place(relx=0.05, rely=0.1, relwidth=0.9, relheight=0.7)

# Day
def process_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet_name_lst = ["兴咸路", "创新港", "天福和园", "沣润和园", "大王镇", "同文路"]
    if '分类清单' in workbook.sheetnames:
        workbook.remove(workbook['分类清单'])
    sheetC = workbook.create_sheet("分类清单")
    category_lst = []
    category_idx = []

    for item in sheet_name_lst:
        if item not in workbook.sheetnames:
            continue
        index = 2
        sheeti = workbook[item]
        while True:
            row = sheeti[index]
            commodity_name = str(row[1].value).strip()
            # print(commodity_name)
            if not commodity_name or commodity_name == "#" or commodity_name == "None":  # 如果品名为空，则认为是数据截止的位置
                break
            commodity_unit = str(row[3].value).strip()
            try:
                commodity_num = float(row[4].value)
            except ValueError:
                commodity_num = row[4].value
            format, price, category = searchInfo(commodity_name, commodity_unit)

            # print(f"data {commodity_name}- {commodity_num}- {commodity_unit}: {format}, {price}, {category}")
            if category in category_lst:
                idx = category_lst.index(category)
                sheetC.cell(row=category_idx[idx], column=2 * idx + 1).value = commodity_name
                sheetC.cell(row=category_idx[idx], column=2 * (idx + 1)).value = commodity_num
                category_idx[idx] += 1
            else:
                idx = len(category_lst)
                sheetC.cell(row=1, column=2 * idx + 1).value = category
                sheetC.cell(row=1, column=2 * (idx + 1)).value = '数量'

                sheetC.cell(row=2, column=2 * idx + 1).value = commodity_name
                sheetC.cell(row=2, column=2 * (idx + 1)).value = commodity_num

                category_lst.append(category)
                category_idx.append(3)

            if format:
                # print(f"data {commodity_name}- {commodity_num}- {commodity_unit}: {format}, {price}, {category}")
                sheeti.cell(row=index, column=3).value = format
            sheeti.cell(row=index, column=7).value = price
            if price:
                sheeti.cell(row=index, column=6).value = round(float(price) / 0.92, 2)
                if isinstance(commodity_num, float):
                    sheeti.cell(row=index, column=8).value = round(float(price), 2) * commodity_num
            else:
                sheeti.cell(row=index, column=3).value = format
                sheeti.cell(row=index, column=8).value = None
            index = index + 1
    workbook.save(file_path)
    messagebox.showinfo(title="通知", message="·已经完成【品牌规格】、【单价】、【优惠单价】和【金额】写入\n\n·已完成分类生成")
def searchInfo(c_name,c_unit):
    csv_reader = csv.reader(open('datas/配置数据.csv','r'))
    for row in csv_reader:
        name, unit, price, format, category = row[0].strip(), row[1].strip(), row[2].strip(), row[3].strip(), row[4].strip()
        if c_name == name and c_unit == unit:
            # print(c_name,c_unit, "format = " ,format, "price", price, category)
            return format, price, category
    return None, None, None

# Month-1
def parse_data(sheet):
    i = 2
    commodity_dict = {}

    while True:
        row = sheet[i]

        commodity_name, commodity_format, commodity_unit = (str(row[i].value).strip() for i in range(1, 4))
        commodity_num = row[4].value
        if commodity_name == "None" or commodity_name == "#":
            return commodity_dict
        try:
            commodity_num = float(commodity_num)
        except Exception:
            messagebox.showerror(title="错误！", message=f"发生错误：商品{ commodity_name } 的数量为{ commodity_num }，转换成数字过程中出错了，检查一下！")
        commodity_price = float(row[6].value)

        # commodity_dict = {"橘子" : [["斤", 2, 2.6, "规格1"], ['箱', 13, 30, "规格2"]]}
        if commodity_name not in commodity_dict:
            commodity_info = [[commodity_unit, commodity_num, commodity_price, commodity_format]]
            commodity_dict[commodity_name] = commodity_info
        else: # 名称已经保存
            commodity_info = commodity_dict[commodity_name]
            unit_index = [index for index, sublist in enumerate(commodity_info) if sublist[0] == commodity_unit]
            if len(unit_index) == 0:
                commodity_info.append([commodity_unit, commodity_num, commodity_price, commodity_format])
            else:
                price_index = next((i for i in unit_index if commodity_info[i][2] == commodity_price), -1)
                if price_index != -1:
                    commodity_info[price_index][1] += commodity_num
                    if commodity_format != commodity_info[price_index][3] and commodity_format != "None":
                        commodity_info[price_index][3] = commodity_format
                else:        # 找不到价格
                    commodity_info.append([commodity_unit, commodity_num, commodity_price, commodity_format])

        i += 1

    return commodity_dict
def write_to_excel(file_path, commodity_dict):
    try:
        workbook = openpyxl.load_workbook(file_path)
        if '归并后数据' in workbook.sheetnames:
            workbook.remove(workbook['归并后数据'])
        sheet = workbook.create_sheet("归并后数据")
        sheet.append(["xxx 202x 年 xx 月 主副食比价表"])

        sheet.append(["货号", "商品名称", "品牌规格", "发货单位", "发货数量", "配送折后单价（元）", "配送折后金额合计（元）",
                      "超市平均单价（元）", "折后供货平均单价（元）", "最终单价（元）","折扣比价金额（元）"])

        j = 1
        # commodity_dict = {"橘子" : [["斤", 2, 2.6, "规格1"], ['箱', 13, 30, "规格2"]]}
        for commodity_name, commodity_info in commodity_dict.items():
            avg_price = sum(item[2] for item in commodity_info) / len(commodity_info)

            for data_info in commodity_info:
                unit, num, price, format = data_info
                if format == "None":
                    format = ""
                write_data = [j, commodity_name, format, unit, num, price, num * price,"",avg_price,"",""]
                # print("write_data : ", write_data)
                sheet.append(write_data)
                j += 1

        workbook.save(file_path)
        messagebox.showinfo(title="通知", message="整合完成")
    except PermissionError:
        messagebox.showinfo(title="通知", message="请关闭文件，否则无法写入归并后的新数据")
    except Exception as e:
        messagebox.showerror(title="错误！", message=f"发生错误：{str(e)}")


# show bg gif
def get_random_gif(pics_folder):
    # 获取指定文件夹下所有 GIF 文件
    gif_files = [f for f in os.listdir(pics_folder) if f.endswith('.gif')]
    # 随机选择一个 GIF 文件
    random_gif = random.choice(gif_files)
    return random_gif

if __name__ == "__main__":
    windows = tk.Tk()
    app = FileOperationsApp(windows)
    windows.mainloop()
